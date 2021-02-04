import os
import pickle
import re
import shutil
import sys

import pyautogui
import pyperclip
import taxjar
import win32com.client
from PySide2.QtCore import *
from PySide2.QtUiTools import *
from PySide2.QtWidgets import *
from openpyxl import load_workbook
from uszipcode import SearchEngine
import re



class MySearch( QWidget ):
    def __init__(self):
        QWidget.__init__( self )
        layout = QFormLayout()

        self.searchBy = QComboBox()

        self.keyword = QLineEdit()

        self.list = QListWidget()

        self.label = QLabel( "Double click job number from list to copy it to your clipboard" )
        print()

        self.searchBy.addItems( ['jobnumber', 'Job Name', 'Salesman', 'Designer', 'Region', 'Street Name',
                                 'Zip Code', 'Quoted Price', 'Total Price', 'Customer Code', 'Customer Name',
                                 'Billing Street', 'Billing City', 'Billing Zip', 'Bf', 'phonenumber',
                                 'email', 'date', 'po', 'payment'] )

        self.btn = QPushButton( "Search" )

        self.btn.clicked.connect( self.searcher )
        self.list.itemDoubleClicked.connect( self.copytoboard )

        layout.addRow( self.keyword )
        layout.addRow( self.searchBy, self.btn )
        layout.addRow( self.label )
        layout.addWidget( self.list )

        self.setLayout( layout )
        self.setWindowTitle( "Search" )

    def searcher(self):
        self.list.clear()

        wd = os.getcwd()
        print( wd )
        jobs = wd + '\\Jobs\\'
        jobs = os.listdir( jobs )
        print( jobs )

        for f in jobs:
            job = pickle.load( open( wd + '\\Jobs\\' + f, 'rb' ) )
            allResults = job.get( self.searchBy.currentText(), '' )
            allResults = allResults.lower()
            key = self.keyword.text()
            key = key.lower()
            if key in str( allResults ):
                print( 'true' )
                self.list.addItem( job.get( 'jobnumber', '' ) )

    def copytoboard(self):
        jn = self.list.currentItem().text()
        pyperclip.copy( jn )


class MyPopup( QWidget ):
    def __init__(self):
        QWidget.__init__( self )
        layout = QFormLayout()
        self.btn = QPushButton( "Remove salesmen" )
        self.btn.clicked.connect( self.remove )
        wd = os.getcwd()

        salesmanList = pickle.load( open( wd + '\\Salesmen\\sales', "rb" ) )

        salesmen = salesmanList.get( 'salesmen' )

        self.le = QComboBox()
        self.le.addItems( salesmen )
        layout.addRow( self.btn, self.le )
        self.btn1 = QPushButton( "get name" )

        self.setLayout( layout )
        self.setWindowTitle( "Input Dialog demo" )

    def remove(self):
        password = pyautogui.prompt( 'Password' )
        if password == '0206':
            salesmenToRemove = self.le.currentText()
            wd = os.getcwd()

            salesmanList = pickle.load( open( wd + '\\Salesmen\\sales', "rb" ) )

            salesmen = salesmanList.get( 'salesmen' )

            salesmen.remove( salesmenToRemove )

            salesman = {'salesmen': salesmen}
            pickle.dump( salesman, open( wd + '\\Salesmen\\sales', "wb" ) )

            pyautogui.alert(
                'You will have to restart the program for these changes to take effect Or you can go to File -> Reload' )
            salesmanList = pickle.load( open( wd + '\\Salesmen\\sales', "rb" ) )
            self.le.clear()
            self.le.addItems( salesmen )

            self.w.close()
        else:
            pyautogui.alert(
                'Sorry, Password incorrect, please contant Admin Paul Sfalanga to remove salesman (352)-460-5117' )


class Form( QObject ):
    # converting qt main file to python so python can edit and get information from it

    def __init__(self, ui_file, parent=None):
        super( Form, self ).__init__( parent )
        ui_file = QFile( ui_file )
        ui_file.open( QFile.ReadOnly )

        loader = QUiLoader()
        self.window = loader.load( ui_file )
        ui_file.close()
        wd = os.getcwd()

        # salesman = {'salesmen':['','Paul Sfalanga III','Monte Harris' ,'Brett Smith','Parker Beaty' ,'Doug Beaty','FM','Rich Huffer']}
        # pickle.dump( salesman, open( wd+'\\Salesmen\\sales', "wb") )

        salesmanList = pickle.load( open( wd + '\\Salesmen\\sales', "rb" ) )
        client = taxjar.Client( api_key='b945f7cd7180bf57e187f91e84cf76e0' )
        salesmen = salesmanList.get( 'salesmen' )
        designers = ['', 'Paul Sfalanga III', 'Monte Harris', 'Brett Smith', 'Parker Beaty', 'Doug Beaty', 'FM',
                     'Rich Huffer', '']

        self.inputJobName = self.window.findChild( QLineEdit, 'inputJobName' )
        self.inputSalesman = self.window.findChild( QComboBox, 'inputSalesman' )
        self.inputSalesman.addItems( salesmen )
        self.inputDesigner = self.window.findChild( QComboBox, 'inputDesigner' )
        self.inputDesigner.addItems( designers )
        self.inputRegion = self.window.findChild( QLineEdit, 'inputRegion' )
        self.inputStreetName = self.window.findChild( QLineEdit, 'inputStreetName' )
        self.inputZipCode = self.window.findChild( QLineEdit, 'inputZipCode' )
        self.inputQuotedPrice = self.window.findChild( QLineEdit, 'inputQuotedPrice' )
        self.inputTotalPrice = self.window.findChild( QLineEdit, 'inputTotalPrice' )
        self.inputCustomerCode = self.window.findChild( QLineEdit, 'inputCustomerCode' )
        self.inputCustomerName = self.window.findChild( QLineEdit, 'inputCustomerName' )
        self.inputBillingStreet = self.window.findChild( QLineEdit, 'inputBillingStreet' )
        self.inputBillingCity = self.window.findChild( QLineEdit, 'inputBillingCity' )
        self.inputBillingZip = self.window.findChild( QLineEdit, 'inputBillingZip' )
        self.inputBf = self.window.findChild( QLineEdit, 'inputBf' )
        self.inputPhoneNumber = self.window.findChild( QLineEdit, 'inputPhoneNumber' )
        self.inputEmail = self.window.findChild( QLineEdit, 'inputEmail' )
        self.inputDate = self.window.findChild( QLineEdit, 'inputDate' )
        self.trussNumber = self.window.findChild( QLineEdit, 'trussNumber' )
        self.inputPo = self.window.findChild( QLineEdit, 'inputPo' )
        self.jobNameEst = self.window.findChild( QLineEdit, 'jobNameEst' )
        self.runNameEst = self.window.findChild( QLineEdit, 'runNameEst' )
        self.trussNumberFloor = self.window.findChild( QLineEdit, 'trussNumberFloor' )
        self.shippingCity = self.window.findChild( QComboBox, 'shippingCity' )

        self.lineEditCreditApp = self.window.findChild( QLineEdit, 'lineEditCreditApp' )
        self.lineEditCreditApproval = self.window.findChild( QLineEdit, 'lineEditCreditApproval' )
        self.lineEditReceivedPo = self.window.findChild( QLineEdit, 'lineEditReceivedPo' )
        self.lineEditPoSigned = self.window.findChild( QLineEdit, 'lineEditPoSigned' )
        self.lineEditAddVer = self.window.findChild( QLineEdit, 'lineEditAddVer' )
        self.lineEditPlansRequested = self.window.findChild( QLineEdit, 'lineEditPlansRequested' )
        self.lineEditPlansSentToDesign = self.window.findChild( QLineEdit, 'lineEditPlansSentToDesign' )
        self.lineEditCadReq = self.window.findChild( QLineEdit, 'lineEditCadReq' )
        self.lineEditCadSent = self.window.findChild( QLineEdit, 'lineEditCadSent' )
        self.lineEditFlrSub = self.window.findChild( QLineEdit, 'lineEditFlrSub' )
        self.lineEditFlrApp = self.window.findChild( QLineEdit, 'lineEditFlrApp' )
        self.lineEditRfSub = self.window.findChild( QLineEdit, 'lineEditRfSub' )
        self.lineEditRfApp = self.window.findChild( QLineEdit, 'lineEditRfApp' )
        self.lineEditBeam = self.window.findChild( QLineEdit, 'lineEditBeam' )
        self.lineEditHang = self.window.findChild( QLineEdit, 'lineEditHang' )
        self.lineEditLay = self.window.findChild( QLineEdit, 'lineEditLay' )
        self.inputRegionCusTax = self.window.findChild( QLineEdit, 'inputRegionCusTax' )

        self.textEditJobNotes = self.window.findChild( QTextEdit, 'textEditJobNotes' )

        self.treeWidget = self.window.findChild( QTreeWidget, 'treeWidget' )

        # Combo Boxes
        self.inputJobNumber = self.window.findChild( QComboBox, 'inputJobNumber' )
        self.pitch = self.window.findChild( QComboBox, 'pitch' )
        self.bcSize = self.window.findChild( QComboBox, 'bcSize' )
        self.depthFloor = self.window.findChild( QComboBox, 'depthFloor' )

        # Spin Boxes
        self.span = self.window.findChild( QSpinBox, 'span' )
        self.gable = self.window.findChild( QSpinBox, 'gable' )
        self.pPerBf = self.window.findChild( QDoubleSpinBox, 'pPerBf' )
        self.spanFloor = self.window.findChild( QSpinBox, 'spanFloor' )

        # QLabels
        self.lPriceBf = self.window.findChild( QLabel, 'lPriceBf' )
        self.lPrice = self.window.findChild( QLabel, 'lPrice' )
        self.lTotalBf = self.window.findChild( QLabel, 'lTotalBf' )
        self.taxRate = self.window.findChild( QLabel, 'taxRate' )

        # CheckBoxes
        self.taxFree = self.window.findChild( QCheckBox, 'taxFree' )
        self.checkBoxCusTax = self.window.findChild( QCheckBox, 'checkBoxCusTax' )

        # Radio Buttons
        self.radioAcccount = self.window.findChild( QRadioButton, 'radioAccount' )
        self.radioCard = self.window.findChild( QRadioButton, 'radioCard' )
        self.radioCheck = self.window.findChild( QRadioButton, 'radioCheck' )
        self.radioCash = self.window.findChild( QRadioButton, 'radioCash' )
        self.comboCust = self.window.findChild( QComboBox, 'comboCust' )

        # Table
        self.tableJobEst = self.window.findChild( QTableWidget, 'tableJobEst' )

        # Buttons
        self.actionRemove_Salesmen = self.window.findChild( QAction, 'actionRemove_Salesmen' )
        self.actionSalesmen = self.window.findChild( QAction, 'actionSalesmen' )
        self.actionDesigner = self.window.findChild( QAction, 'actionDesigner' )
        self.deleteJob = self.window.findChild( QPushButton, 'deleteJob' )
        buttonLoad = self.window.findChild( QPushButton, 'buttonLoad' )
        buttonxQuote = self.window.findChild( QPushButton, 'buttonxQuote' )
        buttonNew = self.window.findChild( QPushButton, 'buttonNew' )
        buttonSave = self.window.findChild( QPushButton, 'buttonSave' )
        buttonCopy = self.window.findChild( QPushButton, 'buttonCopy' )
        buttonxl = self.window.findChild( QPushButton, 'buttonxl' )
        buttonXlOpen = self.window.findChild( QPushButton, 'buttonXlOpen' )
        buttonOpenJobFolder = self.window.findChild( QPushButton, 'buttonOpenJobFolder' )
        buttonxApproval = self.window.findChild( QPushButton, 'buttonxApproval' )
        buttonxQuoteForm = self.window.findChild( QPushButton, 'buttonxQuoteForm' )
        buttonxOrder = self.window.findChild( QPushButton, 'buttonxOrder' )
        buttoncalc = self.window.findChild( QPushButton, 'calculate' )
        calculateFloor = self.window.findChild( QPushButton, 'calculateFloor' )
        addToJobEst = self.window.findChild( QPushButton, 'addToJobEst' )
        jobTotalEst = self.window.findChild( QPushButton, 'jobTotalEst' )
        deleteEst = self.window.findChild( QPushButton, 'deleteEst' )
        saveCust = self.window.findChild( QPushButton, 'saveCust' )
        self.actionReload = self.window.findChild( QAction, "actionReload" )
        self.buttonSearch = self.window.findChild( QPushButton, 'buttonSearch' )

        # TODO add feature
        ######################################    disable buttons that are not being used  ###########################################
        # if self.inputJobNumber.currentText() == '':
        #     buttonSave.setDisabled(True)
        #     buttonXlOpen.setDisabled( True )
        #     buttonOpenJobFolder.setDisabled( True )
        #     buttonxApproval.setDisabled( True )
        #     buttonCopy.setDisabled( True )
        #     buttonxl.setDisabled( True )
        #     buttonxQuoteForm.setDisabled( True )
        #     buttonxOrder.setDisabled( True )
        #
        # if self.inputCustomerCode.text() == '':
        #  saveCust.setDisabled( True )

        # Buttons connect (When Pressed)

        buttonLoad.clicked.connect( self.sendEmail )
        deleteEst.clicked.connect( self.deleteSel )
        buttonNew.clicked.connect( self.new )
        buttonSave.clicked.connect( self.save )
        buttonCopy.clicked.connect( self.copyjob )
        buttonxl.clicked.connect( self.xl )
        buttonXlOpen.clicked.connect( self.openXl )
        buttonOpenJobFolder.clicked.connect( self.openJobFolder )
        buttonxApproval.clicked.connect( self.approval )
        buttonxOrder.clicked.connect( self.order )
        buttoncalc.clicked.connect( self.bfEst )
        calculateFloor.clicked.connect( self.bfEstFloor )
        addToJobEst.clicked.connect( self.jobManagment )
        jobTotalEst.clicked.connect( self.jobTotal )
        self.deleteJob.clicked.connect( self.delete )
        buttonxQuote.clicked.connect( self.sendQuote )
        self.inputBf.editingFinished.connect( self.calcBf )
        self.inputQuotedPrice.editingFinished.connect( self.calcBf )
        self.inputQuotedPrice.editingFinished.connect( self.tax )
        self.inputZipCode.editingFinished.connect( self.tax )
        self.jobNameEst.editingFinished.connect( self.findJob )
        buttonxQuoteForm.clicked.connect( self.quoteForm )
        self.taxFree.stateChanged.connect( self.tax )
        self.checkBoxCusTax.stateChanged.connect( self.tax )
        self.inputRegionCusTax.editingFinished.connect( self.tax )
        # self.shippingCity.currentTextChanged.connect(self.tax)
        self.inputStreetName.editingFinished.connect( self.tax )

        saveCust.clicked.connect( self.saveCustAct )
        self.actionSalesmen.triggered.connect( self.addSalesmen )

        self.buttonSearch.clicked.connect( self.search )

        self.actionReload.triggered.connect( self.reloadprog )
        self.actionRemove_Salesmen.triggered.connect( self.removeSalesmen )
        self.treeWidget.itemDoubleClicked.connect( self.openfile )

        cwd = os.getcwd()
        pathcom = cwd + "/Cust"

        files = os.listdir( pathcom )
        for f in files:
            custs = f
            self.comboCust.addItem( custs )
        self.comboCust.setCurrentText( '' )
        self.comboCust.currentTextChanged.connect( self.custLookup )

        wd = os.getcwd()
        files = os.listdir( wd + '\\Jobs' )
        for f in files:
            job = f
            self.inputJobNumber.addItem( job )

        self.inputJobNumber.setCurrentText( '' )

        self.inputJobNumber.activated.connect( self.load )

        self.window.show()


    # loads the pickle file when you type a job Number in the job number input
    def load(self):

        self.inputRegionCusTax.setText( '' )
        self.checkBoxCusTax.setChecked( False )
        self.treeWidget.clear()
        jobNumber = self.inputJobNumber.currentText()
        wd = os.getcwd()
        files = os.listdir( wd + '\\Jobs' )
        for f in files:
            job = f
            self.inputJobNumber.addItem( job )


        try:
            jobInfo = pickle.load( open( wd + '\\Jobs\\' + jobNumber, "rb" ) )
            jobName = jobInfo.get( 'Job Name', '' )
            self.inputJobName.setText( jobName )
            salesman = jobInfo.get( 'Salesman', '' )
            self.inputSalesman.setCurrentText( salesman )
            designer = jobInfo.get( 'Designer', '' )
            self.inputDesigner.setCurrentText( designer )
            region = jobInfo.get( 'Region', '' )
            self.inputRegion.setText( region )
            streetName = jobInfo.get( 'Street Name', '' )
            self.inputStreetName.setText( streetName )
            zipCode = jobInfo.get( 'Zip Code', '' )
            self.inputZipCode.setText( zipCode )
            quotedPrice = jobInfo.get( 'Quoted Price', '' )
            self.inputQuotedPrice.setText( quotedPrice )
            totalPrice = jobInfo.get( 'Total Price', '' )
            self.inputTotalPrice.setText( totalPrice )
            customerCode = jobInfo.get( 'Customer Code', '' )
            self.comboCust.setCurrentText( customerCode )
            customerName = jobInfo.get( 'Customer Name', '' )
            self.inputCustomerCode.setText( customerCode )
            billingStreet = jobInfo.get( 'Billing Street', '' )

            billingCity = jobInfo.get( 'Billing City', '' )
            self.inputBillingStreet.setText( billingStreet )
            billingZip = jobInfo.get( 'Billing Zip', '' )
            self.inputBillingCity.setText( billingCity )
            totalBf = jobInfo.get( 'Bf', '' )

            phoneNumber = jobInfo.get( 'phonenumber', '' )

            email = jobInfo.get( 'email', '' )

            date = jobInfo.get( 'date', '' )

            po = jobInfo.get( 'po', '' )

            payment = jobInfo.get( 'payment', '' )

            creditApp = jobInfo.get( 'creditApp', '' )
            creditApproval = jobInfo.get( 'creditApproval', '' )
            receivedPo = jobInfo.get( 'creditApp', '' )
            poSigned = jobInfo.get( 'poSigned', '' )

            addVer = jobInfo.get( 'addVer', '' )
            plansRequested = jobInfo.get( 'plansRequested', '' )
            plansSentToDesign = jobInfo.get( 'plansSentToDesign', '' )
            cadReq = jobInfo.get( 'cadReq', '' )
            cadSent = jobInfo.get( 'cadSent', '' )
            flrSub = jobInfo.get( 'flrSub', '' )
            flrApp = jobInfo.get( 'flrApp', '' )
            rfSub = jobInfo.get( 'rfSub', '' )
            rfApp = jobInfo.get( 'rfApp', '' )
            beam = jobInfo.get( 'beam', '' )
            hang = jobInfo.get( 'hang', '' )
            lay = jobInfo.get( 'lay', '' )
            jobNotes = jobInfo.get( 'jobNotes', '' )
            taxState = jobInfo.get( 'taxState', '' )
            tax = jobInfo.get( 'tax', '' )

            self.inputCustomerName.setText( customerName )
            self.inputBillingZip.setText( billingZip )
            self.inputBf.setText( totalBf )
            self.inputPhoneNumber.setText( phoneNumber )
            self.inputEmail.setText( email )
            self.inputDate.setText( date )
            self.inputPo.setText( po )
            self.lineEditCreditApp.setText( creditApp )
            self.lineEditCreditApproval.setText( creditApproval )
            self.lineEditReceivedPo.setText( receivedPo )
            self.lineEditPoSigned.setText( poSigned )
            self.lineEditAddVer.setText( addVer )
            self.lineEditPlansRequested.setText( plansRequested )
            self.lineEditPlansSentToDesign.setText( plansSentToDesign )
            self.lineEditCadReq.setText( cadReq )
            self.lineEditCadSent.setText( cadSent )
            self.lineEditFlrSub.setText( flrSub )
            self.lineEditFlrApp.setText( flrApp )
            self.lineEditFlrSub.setText( rfSub )
            self.lineEditFlrApp.setText( rfApp )
            self.lineEditBeam.setText( beam )
            self.lineEditHang.setText( hang )
            self.lineEditLay.setText( lay )
            self.textEditJobNotes.setText( jobNotes )
            # if taxState == 'cust':
            # self.checkBoxCusTax.setChecked(True)
            self.inputRegionCusTax.setText( tax )
            # print(taxState)

            if payment == 'cash':
                self.radioCash.setChecked( True )
            if payment == 'account':
                self.radioAcccount.setChecked( True )
            if payment == 'check':
                self.radioCheck.setChecked( True )
            if payment == 'card':
                self.radioCard.setChecked( True )

            print( 'here' )
            # path = 'O:\Jobs\\' + jobNumber
            #
            # self.load_project_structure( path, self.treeWidget )




        # if no job found
        except FileNotFoundError:

            self.inputJobName.setText( 'No Job Found' )
            self.inputSalesman.setCurrentText( '' )
            self.inputDesigner.setCurrentText( '' )
            self.inputRegion.setText( '' )
            self.inputStreetName.setText( '' )
            self.inputZipCode.setText( '' )
            self.inputQuotedPrice.setText( '' )
            self.inputTotalPrice.setText( '' )
            self.inputCustomerCode.setText( '' )
            self.inputCustomerName.setText( '' )
            self.inputBillingStreet.setText( '' )
            self.inputBillingCity.setText( '' )
            self.inputBillingZip.setText( '' )
            self.inputBf.setText( '' )
            self.comboCust.setCurrentText( '' )
            self.inputPhoneNumber.setText( '' )
            self.inputEmail.setText( '' )
            self.inputDate.setText( '' )
            self.inputPo.setText( '' )
            self.lineEditCreditApp.setText( 'N\A' )
            self.lineEditCreditApproval.setText( 'N\A' )
            self.lineEditReceivedPo.setText( 'N\A' )
            self.lineEditPoSigned.setText( 'N\A' )
            self.lineEditAddVer.setText( 'N\A' )
            self.lineEditPlansRequested.setText( 'N\A' )
            self.lineEditPlansSentToDesign.setText( 'N\A' )
            self.lineEditCadReq.setText( 'N\A' )
            self.lineEditCadSent.setText( 'N\A' )
            self.lineEditFlrSub.setText( 'N\A' )
            self.lineEditFlrApp.setText( 'N\A' )
            self.lineEditFlrSub.setText( 'N\A' )
            self.lineEditFlrApp.setText( 'N\A' )
            self.lineEditBeam.setText( 'N\A' )
            self.lineEditHang.setText( 'N\A' )
            self.lineEditLay.setText( 'N\A' )
            self.textEditJobNotes.setText( 'N\A' )

            self.radioCard.setChecked( False )
            self.radioCheck.setChecked( False )
            self.radioAcccount.setChecked( False )
            self.radioCash.setChecked( False )

            try:
                path = 'O:\Jobs\\' + jobNumber
                icom_file = open(path+'\\JOB.CTL','r')
                icom_info = icom_file.readlines()
                self.inputJobName.setText( re.sub('[|]','',icom_info[3]) )
                self.inputStreetName.setText(re.sub('[|]','',icom_info[4])+' '+re.sub( '[|]', '', icom_info[5] ))


            except:
                pyautogui.alert('Job does not exist in icommand')



        except:
            jobInfo = pickle.load( open( wd + '\\Jobs\\' + jobNumber, "rb" ) )
            jobName = jobInfo.get( 'Job Name', '' )
            self.inputJobName.setText( jobName )
            salesman = jobInfo.get( 'Salesman', '' )
            self.inputSalesman.setCurrentText( salesman )
            designer = jobInfo.get( 'Designer', '' )
            self.inputDesigner.setCurrentText( designer )
            region = jobInfo.get( 'Region', '' )
            self.inputRegion.setText( region )
            streetName = jobInfo.get( 'Street Name', '' )
            self.inputStreetName.setText( streetName )
            zipCode = jobInfo.get( 'Zip Code', '' )
            self.inputZipCode.setText( zipCode )
            quotedPrice = jobInfo.get( 'Quoted Price', '' )
            self.inputQuotedPrice.setText( quotedPrice )
            totalPrice = jobInfo.get( 'Total Price', '' )
            self.inputTotalPrice.setText( totalPrice )
            customerCode = jobInfo.get( 'Customer Code', '' )
            self.comboCust.setCurrentText( customerCode )
            customerName = jobInfo.get( 'Customer Name', '' )
            self.inputCustomerCode.setText( customerCode )
            billingStreet = jobInfo.get( 'Billing Street', '' )

            billingCity = jobInfo.get( 'Billing City', '' )
            self.inputBillingStreet.setText( billingStreet )
            billingZip = jobInfo.get( 'Billing Zip', '' )
            self.inputBillingCity.setText( billingCity )
            totalBf = jobInfo.get( 'Bf', '' )

            phoneNumber = jobInfo.get( 'phonenumber', '' )

            email = jobInfo.get( 'email', '' )

            date = jobInfo.get( 'date', '' )

            po = jobInfo.get( 'po', '' )

            payment = jobInfo.get( 'payment', '' )

            creditApp = jobInfo.get( 'creditApp', '' )
            creditApproval = jobInfo.get( 'creditApproval', '' )
            receivedPo = jobInfo.get( 'creditApp', '' )
            poSigned = jobInfo.get( 'poSigned', '' )

            addVer = jobInfo.get( 'addVer', '' )
            plansRequested = jobInfo.get( 'plansRequested', '' )
            plansSentToDesign = jobInfo.get( 'plansSentToDesign', '' )
            cadReq = jobInfo.get( 'cadReq', '' )
            cadSent = jobInfo.get( 'cadSent', '' )
            flrSub = jobInfo.get( 'flrSub', '' )
            flrApp = jobInfo.get( 'flrApp', '' )
            rfSub = jobInfo.get( 'rfSub', '' )
            rfApp = jobInfo.get( 'rfApp', '' )
            beam = jobInfo.get( 'beam', '' )
            hang = jobInfo.get( 'hang', '' )
            lay = jobInfo.get( 'lay', '' )
            jobNotes = jobInfo.get( 'jobNotes', '' )
            taxState = jobInfo.get( 'taxState', '' )
            tax = jobInfo.get( 'tax', '' )

            self.inputCustomerName.setText( customerName )
            self.inputBillingZip.setText( billingZip )
            self.inputBf.setText( totalBf )
            self.inputPhoneNumber.setText( phoneNumber )
            self.inputEmail.setText( email )
            self.inputDate.setText( date )
            self.inputPo.setText( po )
            self.lineEditCreditApp.setText( creditApp )
            self.lineEditCreditApproval.setText( creditApproval )
            self.lineEditReceivedPo.setText( receivedPo )
            self.lineEditPoSigned.setText( poSigned )
            self.lineEditAddVer.setText( addVer )
            self.lineEditPlansRequested.setText( plansRequested )
            self.lineEditPlansSentToDesign.setText( plansSentToDesign )
            self.lineEditCadReq.setText( cadReq )
            self.lineEditCadSent.setText( cadSent )
            self.lineEditFlrSub.setText( flrSub )
            self.lineEditFlrApp.setText( flrApp )
            self.lineEditFlrSub.setText( rfSub )
            self.lineEditFlrApp.setText( rfApp )
            self.lineEditBeam.setText( beam )
            self.lineEditHang.setText( hang )
            self.lineEditLay.setText( lay )
            self.textEditJobNotes.setText( jobNotes )
            # if taxState == 'cust':
            # self.checkBoxCusTax.setChecked(True)
            self.inputRegionCusTax.setText( tax )
            # print(taxState)

            if payment == 'cash':
                self.radioCash.setChecked( True )
            if payment == 'account':
                self.radioAcccount.setChecked( True )
            if payment == 'check':
                self.radioCheck.setChecked( True )
            if payment == 'card':
                self.radioCard.setChecked( True )

            print( 'here' )



    # runs jobdat.py which gets the data from icommand
    def new(self):
        from disgrarded import jobdat
        jobdat.new()

    # saves info to a pickle file
    def save(self):
        jobNumber = self.inputJobNumber.currentText()
        jobName = self.inputJobName.text()
        salesman = self.inputSalesman.currentText()
        designer = self.inputDesigner.currentText()
        region = self.inputRegion.text()
        street = self.inputStreetName.text()
        zipCode = self.inputZipCode.text()
        quotedPrice = self.inputQuotedPrice.text()
        totalPrice = self.inputTotalPrice.text()
        customerCode = self.inputCustomerCode.text()
        customerName = self.inputCustomerName.text()
        billingStreet = self.inputBillingStreet.text()
        billingCity = self.inputBillingCity.text()
        billingZip = self.inputBillingZip.text()
        totalBf = self.inputBf.text()
        phoneNumber = self.inputPhoneNumber.text()
        email = self.inputEmail.text()
        date = self.inputDate.text()
        po = self.inputPo.text()
        creditApp = self.lineEditCreditApp.text()
        creditApproval = self.lineEditCreditApproval.text()
        receivedPo = self.lineEditReceivedPo.text()
        poSigned = self.lineEditPoSigned.text()
        addVer = self.lineEditAddVer.text()
        plansRequested = self.lineEditPlansRequested.text()
        plansSentToDesign = self.lineEditPlansSentToDesign.text()
        cadReq = self.lineEditCadReq.text()
        cadSent = self.lineEditCadSent.text()
        flrSub = self.lineEditFlrSub.text()
        flrApp = self.lineEditFlrApp.text()
        rfSub = self.lineEditRfSub.text()
        rfApp = self.lineEditRfApp.text()
        beam = self.lineEditBeam.text()
        hang = self.lineEditHang.text()
        lay = self.lineEditLay.text()
        jobNotes = self.textEditJobNotes.toPlainText()

        tax = self.inputRegionCusTax.text()
        if self.checkBoxCusTax.isChecked():
            taxState = 'cust'
        else:
            taxState = 'nonCust'

        if self.radioCash.isChecked():
            payment = 'cash'
        if self.radioAcccount.isChecked():
            payment = 'account'
        if self.radioCard.isChecked():
            payment = 'card'
        if self.radioCheck.isChecked():
            payment = 'check'


        try:

            jobInfo = {'jobnumber': jobNumber, 'Job Name': jobName, 'Salesman': salesman, 'Designer': designer,
                       'Region': region, 'Street Name': street,
                       'Zip Code': zipCode, 'Quoted Price': quotedPrice, 'Total Price': totalPrice,
                       'Customer Code': customerCode, 'Customer Name': customerName,
                       'Billing Street': billingStreet, 'Billing City': billingCity, 'Billing Zip': billingZip,
                       'Bf': totalBf, 'phonenumber': phoneNumber,
                       'email': email, 'date': date, 'po': po, 'payment': payment, 'creditApp': creditApp,
                       'creditApproval': creditApproval, 'receivedPo': receivedPo, 'poSigned': poSigned, 'addVer': addVer
                , 'plansRequested': plansRequested, 'plansSentToDesign': plansSentToDesign, 'cadReq': cadReq,
                       'cadSent': cadSent, 'flrSub': flrSub, 'flrApp': flrApp, 'rfSub': rfSub, 'rfApp': rfApp
                , 'beam': beam, 'hang': hang, 'lay': lay, 'jobNotes': jobNotes, 'tax': tax, 'taxState': taxState}
            pickle.dump( jobInfo, open( jobNumber, "wb" ) )

            file = jobNumber
            shutil.move( os.getcwd() + '\\' + file, os.getcwd() + '\\Jobs\\' + file )
            pyautogui.alert('job saved')
        except:

            pyautogui.alert("Unknown payment type! please select paymet type.")


    # creates and xl file to print pdfs from
    def xl(self):

        try:
            jobNumber = self.inputJobNumber.currentText()
            jobName = self.inputJobName.text()
            salesman = self.inputSalesman.currentText()
            designer = self.inputDesigner.currentText()
            region = self.inputRegion.text()
            street = self.inputStreetName.text()
            zipCode = self.inputZipCode.text()
            quotedPrice = self.inputQuotedPrice.text()
            totalPrice = self.inputTotalPrice.text()
            customerCode = self.inputCustomerCode.text()
            customerName = self.inputCustomerName.text()
            billingStreet = self.inputBillingStreet.text()
            billingCity = self.inputBillingCity.text()
            billingZip = self.inputBillingZip.text()
            totalBf = self.inputBf.text()
            phoneNumber = self.inputPhoneNumber.text()
            email = self.inputEmail.text()
            date = self.inputDate.text()
            po = self.inputPo.text()

            creditApp = self.lineEditCreditApp.text()
            creditApproval = self.lineEditCreditApproval.text()
            receivedPo = self.lineEditReceivedPo.text()
            poSigned = self.lineEditPoSigned.text()
            addVer = self.lineEditAddVer.text()
            plansRequested = self.lineEditPlansRequested.text()
            plansSentToDesign = self.lineEditPlansSentToDesign.text()
            cadReq = self.lineEditCadReq.text()
            cadSent = self.lineEditCadSent.text()
            flrSub = self.lineEditFlrSub.text()
            flrApp = self.lineEditFlrApp.text()
            rfSub = self.lineEditRfSub.text()
            rfApp = self.lineEditRfApp.text()
            beam = self.lineEditBeam.text()
            hang = self.lineEditHang.text()
            lay = self.lineEditLay.text()
            jobNotes = self.textEditJobNotes.toPlainText()

            tax = self.inputRegionCusTax.text()

            if self.radioCash.isChecked():
                payment = 'cash'
            if self.radioAcccount.isChecked():
                payment = 'account'
            if self.radioCard.isChecked():
                payment = 'card'
            if self.radioCheck.isChecked():
                payment = 'check'

            wb = load_workbook( 'walk.xlsx' )
            ws = wb['Entry']
            ws2 = wb['Order Form']
            ws['C4'] = jobNumber
            ws['D4'] = customerCode
            ws['B4'] = jobName
            ws['E4'] = salesman
            ws['F4'] = salesman
            ws['G4'] = phoneNumber
            ws['B16'] = date
            ws['H4'] = email
            # TODO add state
            ws['I4'] = billingStreet + ", " + billingCity + ', ' + billingZip
            if self.taxFree.isChecked():
                ws['J10'] = True
                pyautogui.alert( 'Warning: No Tax applied ' )

            ws['B10'] = totalBf
            ws['C10'] = quotedPrice
            ws['F10'] = float( zipCode )
            ws['C16'] = street
            ws['B19'] = customerName
            ws['C19'] = phoneNumber
            ws['F16'] = po
            if self.checkBoxCusTax.isChecked():
                tax = self.inputRegionCusTax.text()
                ws2['B28'] = str( tax )

            ws2['A7'] = creditApp
            ws2['B7'] = creditApproval
            ws2['D7'] = receivedPo
            ws2['F7'] = poSigned
            ws2['H7'] = addVer
            ws2['A9'] = plansRequested
            ws2['B9'] = plansSentToDesign
            ws2['D9'] = cadReq
            ws2['F9'] = cadSent
            ws2['H9'] = flrSub
            ws2['A11'] = flrApp
            ws2['B11'] = rfSub
            ws2['D11'] = rfApp
            ws2['F11'] = beam
            ws2['H11'] = hang
            ws2['A13'] = lay
            ws2['A41'] = jobNotes
            ws2['B28'] = str( tax ) + "%"
            if payment == 'cash':
                ws2['B31'] = 'X'
            if payment == 'check':
                ws2['D31'] = 'X'
            if payment == 'account':
                ws2['H31'] = 'X'
            if payment == 'card':
                ws2['F31'] = 'X'
            wd = os.getcwd()

            wb.save( wd + '\\xldocs\\' + jobNumber + '.xlsx' )
            pyautogui.alert( 'Data Consolidate' )
            viewXL = pyautogui.confirm( text='Would you like to view the excel now?', title='Veiw XL',
                                        buttons=['Yes', 'No'] )

            if viewXL == 'Yes':
                self.openXl()
        except:
            pyautogui.alert(
                'Data Consolidate Failed. Insure excel is closed and retry. If it still fails try to restart the program. For help please contact Paul at 352-460-5117' )

    # Creates pdf approval form and puts it in the job folder
    def approval(self):
        try:
            import win32com.client

            jobNumber = self.inputJobNumber.currentText()

            o = win32com.client.Dispatch( "Excel.Application" )

            o.Visible = True

            wb_path = os.getcwd() + '\\xldocs\\' + jobNumber + '.xlsx'

            wb = o.Workbooks.Open( wb_path )

            ws_index_list = [3]  # say you want to print these sheets

            path_to_pdf = os.getcwd() + '\\xldocs\\' + jobNumber + ' Approval.pdf'

            print_area = 'A1:L50'

            for index in ws_index_list:

                # off-by-one so the user can start numbering the worksheets at 1

                ws = wb.Worksheets[index - 1]

                ws.PageSetup.Zoom = False

                ws.PageSetup.FitToPagesTall = 1

                ws.PageSetup.FitToPagesWide = 1

                ws.PageSetup.PrintArea = print_area

            wb.WorkSheets( ws_index_list ).Select()

            wb.ActiveSheet.ExportAsFixedFormat( 0, path_to_pdf )
            wb.Close( True )
            pyautogui.alert( 'Approval Created' )

            if not os.path.exists( r'O:\Jobs\\' + jobNumber + '\Orders' ):
                os.mkdir( r'O:\Jobs\\' + jobNumber + '\Orders' )

            file = jobNumber + ' Approval.pdf'
            shutil.move( os.getcwd() + '\\xldocs\\' + file, 'O:\Jobs\\' + jobNumber + '\Orders\\' + file )




        except:
            pyautogui.alert( 'Unable to creatE pdf Have you Consolidated this job yet?' )

    # Creates pdf approval form and puts it in the job folder
    def order(self):

        try:

            jobNumber = self.inputJobNumber.currentText()

            o = win32com.client.Dispatch( "Excel.Application" )

            o.Visible = False

            wb_path = os.getcwd() + '\\xldocs\\' + jobNumber + '.xlsx'

            wb = o.Workbooks.Open( wb_path )

            ws_index_list = [4]  # say you want to print these sheets

            path_to_pdf = os.getcwd() + '\\xldocs\\' + jobNumber + ' Order.pdf'

            print_area = 'A1:J50'

            for index in ws_index_list:

                # off-by-one so the user can start numbering the worksheets at 1

                ws = wb.Worksheets[index - 1]

                ws.PageSetup.Zoom = False

                ws.PageSetup.FitToPagesTall = 1

                ws.PageSetup.FitToPagesWide = 1

                ws.PageSetup.PrintArea = print_area

            wb.WorkSheets( ws_index_list ).Select()

            wb.ActiveSheet.ExportAsFixedFormat( 0, path_to_pdf )
            wb.Close( True )
            pyautogui.alert( 'Order Form Created' )
            if not os.path.exists( r'O:\Jobs\\' + jobNumber + '\Orders' ):
                os.mkdir( r'O:\Jobs\\' + jobNumber + '\Orders' )

            file = jobNumber + ' Order.pdf'
            shutil.move( os.getcwd() + '\\xldocs\\' + file, 'O:\Jobs\\' + jobNumber + '\Orders\\' + file )

        except:
            pyautogui.alert( 'Unable to create pdf Have you Consolidated this job yet?' )

    # Creates email to send to Production using outlook
    def sendEmail(self):
        jobNumber = self.inputJobNumber.currentText()
        jobName = self.inputJobName.text()
        phoneNumber = self.inputPhoneNumber.text()
        date = self.inputDate.text()
        customerName = self.inputCustomerName.text()
        hangers = self.lineEditHang.text()
        beams = self.lineEditBeam.text()
        jobNotes = self.textEditJobNotes.toPlainText()

        sub = str( jobNumber ) + '-' + str( jobName ) + '-' + 'Order'
        const = win32com.client.constants
        olMailItem = 0x0
        obj = win32com.client.Dispatch( "Outlook.Application" )
        newMail = obj.CreateItem( olMailItem )
        newMail.Subject = sub
        newMail.BodyFormat = 2  # olFormatHTML https://msdn.microsoft.com/en-us/library/office/aa219371(v=office.11).aspx
        newMail.HTMLBody = "<HTML><BODY>This one is ready for production <br><br> Thanks,<br><br>Delivery Date:" + date + "<br>Call before Delivery " + phoneNumber + "<br><br>Hangers Ordered: " + hangers + "<br> Beams Ordered: " + beams + "<br> Special Instructions: " + jobNotes + " <br><br> Paul Sfalanga III<br>(864)772-3423</BODY></HTML>"
        newMail.To = "tstrayer@paneltruss.com; ty@paneltruss.com; mlowe@paneltruss.com; dickie@paneltruss.com; amarsingill@paneltruss.com; dlawrence@paneltruss.com; akimsey@paneltruss.com"

        newMail.display()
        # newMail.Send()

    # Creates email to send to customer using outlook
    def sendQuote(self):

        jobNumber = self.inputJobNumber.currentText()
        jobName = self.inputJobName.text()
        phoneNumber = self.inputPhoneNumber.text()
        emailAddress = self.inputEmail.text()
        date = self.inputDate.text()
        customerName = self.inputCustomerName
        sub = str( jobNumber ) + '-' + str( jobName ) + '-' + ' Panel Truss Quote'
        const = win32com.client.constants
        olMailItem = 0x0
        obj = win32com.client.Dispatch( "Outlook.Application" )
        newMail = obj.CreateItem( olMailItem )
        newMail.Subject = sub
        newMail.BodyFormat = 2  # olFormatHTML https://msdn.microsoft.com/en-us/library/office/aa219371(v=office.11).aspx
        newMail.HTMLBody = "<HTML><BODY>Attached to this email it the quote for the trusses you requested.<br><br>Let me know if you have any questions <br>Thanks,<br><br>Paul Sfalanga III<br>(864)772-3423</BODY></HTML>"
        newMail.To = emailAddress

        newMail.display()

    # fills the price pre bf label with the correct price/ BF
    def calcBf(self):
        bf = float( self.inputBf.text() )
        quotedPrice = self.inputQuotedPrice.text()
        quotedPrice = re.sub( '[!@#$,]', '', quotedPrice )

        print( quotedPrice )
        quotedPrice = float( quotedPrice )
        bfPrice = quotedPrice / bf
        bfPrice = round( bfPrice, 2 )
        bfPrice = 'Price/BF: $' + str( bfPrice )
        self.lPriceBf.setText( bfPrice )

    def bfEst(self):

        pitch = str( self.pitch.currentText() )
        span = str( self.span.value() )
        gable = str( self.gable.value() )
        trussNumber = self.trussNumber.text()
        pPerBf = self.pPerBf.value()
        bfCode = 'p' + pitch + 's' + span

        if self.bcSize.currentText() == '2x4':
            bfCalcDict = {
                'p3/12s10': 15.3333,
                'p3/12s11': 17.3333,
                'p3/12s12': 18.6667,
                'p3/12s13': 20,
                'p3/12s14': 21.3333,
                'p3/12s15': 22.6667,
                'p3/12s16': 25.3333,
                'p3/12s17': 30.64,
                'p3/12s18': 33.3333,
                'p3/12s19': 34.6667,
                'p3/12s20': 37.3333,
                'p3/12s21': 40,
                'p3/12s22': 41.3333,
                'p3/12s23': 42.6667,
                'p3/12s24': 45.3333,
                'p3/12s25': 48.6667,
                'p3/12s26': 48.6667,
                'p3/12s27': 52,
                'p3/12s28': 54.6667,
                'p3/12s29': 56,
                'p3/12s30': 56,
                'p3/12s31': 60.6667,
                'p3/12s32': 60,
                'p3/12s33': 68,
                'p3/12s34': 70.6667,
                'p3/12s35': 74.6667,
                'p3/12s36': 86.6667,
                'p3/12s37': 88.6667,
                'p3/12s38': 88.6667,
                'p3/12s39': 100.6667,
                'p3/12s40': 99.3333,
                'p4/12s10': 16,
                'p4/12s11': 17.3333,
                'p4/12s12': 18.6667,
                'p4/12s13': 20,
                'p4/12s14': 22,
                'p4/12s15': 26,
                'p4/12s16': 26,
                'p4/12s17': 33.3333,
                'p4/12s18': 34.6667,
                'p4/12s19': 40,
                'p4/12s20': 40,
                'p4/12s21': 41.3333,
                'p4/12s22': 41.3333,
                'p4/12s23': 46.6667,
                'p4/12s24': 46.6667,
                'p4/12s25': 49.3333,
                'p4/12s26': 52.6667,
                'p4/12s27': 56.6667,
                'p4/12s28': 56.6667,
                'p4/12s29': 58,
                'p4/12s30': 59.3333,
                'p4/12s31': 61.3333,
                'p4/12s32': 69.3333,
                'p4/12s33': 74.6667,
                'p4/12s34': 77.3333,
                'p4/12s35': 78.6667,
                'p4/12s36': 78.6667,
                'p4/12s37': 81.3333,
                'p4/12s38': 85.3333,
                'p4/12s39': 86.6667,
                'p4/12s40': 91.3333,
                'p5/12s10': 16,
                'p5/12s11': 19.3333,
                'p5/12s12': 19.3333,
                'p5/12s13': 22,
                'p5/12s14': 22,
                'p5/12s15': 26,
                'p5/12s16': 26.6667,
                'p5/12s17': 36,
                'p5/12s18': 36,
                'p5/12s19': 40,
                'p5/12s20': 40,
                'p5/12s21': 42.6667,
                'p5/12s22': 45.3333,
                'p5/12s23': 46.6667,
                'p5/12s24': 50,
                'p5/12s25': 52.6667,
                'p5/12s26': 56,
                'p5/12s27': 57.3333,
                'p5/12s28': 57.3333,
                'p5/12s29': 60,
                'p5/12s30': 62.6667,
                'p5/12s31': 64,
                'p5/12s32': 74.6667,
                'p5/12s33': 77.3333,
                'p5/12s34': 81.3333,
                'p5/12s35': 82.6667,
                'p5/12s36': 85.3333,
                'p5/12s37': 89.3333,
                'p5/12s38': 90.6667,
                'p5/12s39': 93.3333,
                'p5/12s40': 97.3333,
                'p6/12s10': 16.6667,
                'p6/12s11': 19.3333,
                'p6/12s12': 19.3333,
                'p6/12s13': 22,
                'p6/12s14': 22.6667,
                'p6/12s15': 26.6667,
                'p6/12s16': 26.6667,
                'p6/12s17': 36,
                'p6/12s18': 40,
                'p6/12s19': 41.3333,
                'p6/12s20': 41.3333,
                'p6/12s21': 44,
                'p6/12s22': 48,
                'p6/12s23': 49.3333,
                'p6/12s24': 52,
                'p6/12s25': 57.3333,
                'p6/12s26': 58,
                'p6/12s27': 60.6667,
                'p6/12s28': 60.6667,
                'p6/12s29': 63.3333,
                'p6/12s30': 64,
                'p6/12s31': 70,
                'p6/12s32': 82.6667,
                'p6/12s33': 84,
                'p6/12s34': 85.3333,
                'p6/12s35': 89.3333,
                'p6/12s36': 92,
                'p6/12s37': 93.3333,
                'p6/12s38': 93.3333,
                'p6/12s39': 100,
                'p6/12s40': 104,
                'p7/12s10': 16.6667,
                'p7/12s11': 19.3333,
                'p7/12s12': 21.3333,
                'p7/12s13': 22.6667,
                'p7/12s14': 25.3333,
                'p7/12s15': 27.3333,
                'p7/12s16': 36,
                'p7/12s17': 40,
                'p7/12s18': 40,
                'p7/12s19': 44,
                'p7/12s20': 44,
                'p7/12s21': 48,
                'p7/12s22': 49.3333,
                'p7/12s23': 50.6667,
                'p7/12s24': 56.6667,
                'p7/12s25': 58.6667,
                'p7/12s26': 60,
                'p7/12s27': 61.3333,
                'p7/12s28': 64,
                'p7/12s29': 66.6667,
                'p7/12s30': 72,
                'p7/12s31': 86.6667,
                'p7/12s32': 86.6667,
                'p7/12s33': 88,
                'p7/12s34': 90.6667,
                'p7/12s35': 98.6667,
                'p7/12s36': 98.6667,
                'p7/12s37': 100,
                'p7/12s38': 102.6667,
                'p7/12s39': 106.6667,
                'p7/12s40': 106.6667,
                'p8/12s10': 18.6667,
                'p8/12s11': 20,
                'p8/12s12': 21.3333,
                'p8/12s13': 26,
                'p8/12s14': 26,
                'p8/12s15': 27.3333,
                'p8/12s16': 36,
                'p8/12s17': 42.6667,
                'p8/12s18': 42.6667,
                'p8/12s19': 44,
                'p8/12s20': 48,
                'p8/12s21': 49.3333,
                'p8/12s22': 49.3333,
                'p8/12s23': 57.3333,
                'p8/12s24': 57.3333,
                'p8/12s25': 62.6667,
                'p8/12s26': 62.6667,
                'p8/12s27': 65.3333,
                'p8/12s28': 68,
                'p8/12s29': 70.6667,
                'p8/12s30': 73.3333,
                'p8/12s31': 90.6667,
                'p8/12s32': 90.6667,
                'p8/12s33': 97.3333,
                'p8/12s34': 97.3333,
                'p8/12s35': 102.6667,
                'p8/12s36': 117.3333,
                'p8/12s37': 121.3333,
                'p8/12s38': 124,
                'p8/12s39': 126.6667,
                'p8/12s40': 130.6667,
                'p9/12s10': 18.6667,
                'p9/12s11': 21.3333,
                'p9/12s12': 22,
                'p9/12s13': 26,
                'p9/12s14': 26,
                'p9/12s15': 28,
                'p9/12s16': 41.3333,
                'p9/12s17': 42.6667,
                'p9/12s18': 44,
                'p9/12s19': 48,
                'p9/12s20': 48,
                'p9/12s21': 53.3333,
                'p9/12s22': 53.3333,
                'p9/12s23': 57.3333,
                'p9/12s24': 61.3333,
                'p9/12s25': 62.6667,
                'p9/12s26': 64,
                'p9/12s27': 72,
                'p9/12s28': 72,
                'p9/12s29': 76,
                'p9/12s30': 76,
                'p9/12s31': 92,
                'p9/12s32': 102.6667,
                'p9/12s33': 116,
                'p9/12s34': 118.6667,
                'p9/12s35': 122.6667,
                'p9/12s36': 125.3334,
                'p9/12s37': 129.3334,
                'p9/12s38': 130.6666,
                'p9/12s39': 150,
                'p9/12s40': 154.6666,
                'p10/12s10': 18.6667,
                'p10/12s11': 22,
                'p10/12s12': 24.6667,
                'p10/12s13': 26.6667,
                'p10/12s14': 26.6667,
                'p10/12s15': 30.6667,
                'p10/12s16': 41.3333,
                'p10/12s17': 44,
                'p10/12s18': 44,
                'p10/12s19': 52,
                'p10/12s20': 52,
                'p10/12s21': 53.3333,
                'p10/12s22': 56,
                'p10/12s23': 61.3333,
                'p10/12s24': 61.3333,
                'p10/12s25': 66.6667,
                'p10/12s26': 72,
                'p10/12s27': 72,
                'p10/12s28': 74.6667,
                'p10/12s29': 141.3333,
                'p10/12s30': 98,
                'p10/12s31': 100.6667,
                'p10/12s32': 103.3334,
                'p10/12s33': 107.3333,
                'p10/12s34': 110,
                'p10/12s35': 114,
                'p10/12s36': 130.6667,
                'p10/12s37': 140,
                'p10/12s38': 142.6667,
                'p10/12s39': 145.3333,
                'p10/12s40': 148.6667,
                'p11/12s10': 20.6667,
                'p11/12s11': 22,
                'p11/12s12': 25.3333,
                'p11/12s13': 26.6667,
                'p11/12s14': 27.3333,
                'p11/12s15': 31.3333,
                'p11/12s16': 42.6667,
                'p11/12s17': 48,
                'p11/12s18': 50.6667,
                'p11/12s19': 52,
                'p11/12s20': 52,
                'p11/12s21': 57.3333,
                'p11/12s22': 60,
                'p11/12s23': 61.3333,
                'p11/12s24': 65.3333,
                'p11/12s25': 72,
                'p11/12s26': 70.6667,
                'p11/12s27': 92,
                'p11/12s28': 94.6667,
                'p11/12s29': 98.6667,
                'p11/12s30': 102.6667,
                'p11/12s31': 106.6667,
                'p11/12s32': 109.3334,
                'p11/12s33': 126.6666,
                'p11/12s34': 132.6666,
                'p11/12s35': 139.3334,
                'p11/12s36': 140,
                'p11/12s37': 144,
                'p11/12s38': 150,
                'p11/12s39': 152.6667,
                'p11/12s40': 154.6667,
                'p12/12s10': 20.6667,
                'p12/12s11': 25.3333,
                'p12/12s12': 25.3333,
                'p12/12s13': 27.3333,
                'p12/12s14': 30,
                'p12/12s15': 32,
                'p12/12s16': 46.6667,
                'p12/12s17': 50.6667,
                'p12/12s18': 50.6667,
                'p12/12s19': 52,
                'p12/12s20': 58.6667,
                'p12/12s21': 60,
                'p12/12s22': 60,
                'p12/12s23': 62.6667,
                'p12/12s24': 72,
                'p12/12s25': 86,
                'p12/12s26': 88.6667,
                'p12/12s27': 92.6667,
                'p12/12s28': 96.6667,
                'p12/12s29': 99.3334,
                'p12/12s30': 103.3334,
                'p12/12s31': 121.3334,
                'p12/12s32': 126.6666,
                'p12/12s33': 134,
                'p12/12s34': 136.6667,
                'p12/12s35': 142,
                'p12/12s36': 144,
                'p12/12s37': 147.3333,
                'p12/12s38': 150,
                'p12/12s39': 172.6667,
                'p12/12s40': 188.6667,

            }

        if self.bcSize.currentText() == '2x6':

            bfCalcDict = {

                'p3/12s10': 18,
                'p3/12s11': 20,
                'p3/12s12': 22.6667,
                'p3/12s13': 24.6667,
                'p3/12s14': 26,
                'p3/12s15': 28,
                'p3/12s16': 30.6667,
                'p3/12s17': 38,
                'p3/12s18': 39.3333,
                'p3/12s19': 41.6667,
                'p3/12s20': 44,
                'p3/12s21': 46,
                'p3/12s22': 47.3333,
                'p3/12s23': 50.6667,
                'p3/12s24': 53.3333,
                'p3/12s25': 55.3333,
                'p3/12s26': 55.3333,
                'p3/12s27': 57.3333,
                'p3/12s28': 64,
                'p3/12s29': 66,
                'p3/12s30': 66,
                'p3/12s31': 69.6667,
                'p3/12s32': 72,
                'p3/12s33': 74,
                'p3/12s34': 79.3333,
                'p3/12s35': 84,
                'p3/12s36': 86.6667,
                'p3/12s37': 88.6667,
                'p3/12s38': 88.6667,
                'p3/12s39': 110,
                'p3/12s40': 122,
                'p4/12s10': 18,
                'p4/12s11': 21.3333,
                'p4/12s12': 22.6667,
                'p4/12s13': 24.6667,
                'p4/12s14': 26,
                'p4/12s15': 31.3333,
                'p4/12s16': 31.3333,
                'p4/12s17': 39.3333,
                'p4/12s18': 39.3333,
                'p4/12s19': 45.3333,
                'p4/12s20': 46.6667,
                'p4/12s21': 48.6667,
                'p4/12s22': 48.6667,
                'p4/12s23': 53.3333,
                'p4/12s24': 54.6667,
                'p4/12s25': 56.6667,
                'p4/12s26': 56.6667,
                'p4/12s27': 66,
                'p4/12s28': 66,
                'p4/12s29': 68,
                'p4/12s30': 68,
                'p4/12s31': 73,
                'p4/12s32': 72.6667,
                'p4/12s33': 82,
                'p4/12s34': 86,
                'p4/12s35': 90.6667,
                'p4/12s36': 90.6667,
                'p4/12s37': 92.6667,
                'p4/12s38': 96.6667,
                'p4/12s39': 100,
                'p4/12s40': 100,
                'p5/12s10': 18,
                'p5/12s11': 22.6667,
                'p5/12s12': 23.3333,
                'p5/12s13': 26.6667,
                'p5/12s14': 26.6667,
                'p5/12s15': 31.3333,
                'p5/12s16': 31.3333,
                'p5/12s17': 40.6667,
                'p5/12s18': 42,
                'p5/12s19': 46.6667,
                'p5/12s20': 46.6667,
                'p5/12s21': 50,
                'p5/12s22': 52.6667,
                'p5/12s23': 54.6667,
                'p5/12s24': 54.6667,
                'p5/12s25': 59.3333,
                'p5/12s26': 62,
                'p5/12s27': 66.6667,
                'p5/12s28': 66.6667,
                'p5/12s29': 70,
                'p5/12s30': 71.3333,
                'p5/12s31': 74.3333,
                'p5/12s32': 78,
                'p5/12s33': 85.3333,
                'p5/12s34': 88,
                'p5/12s35': 94.6667,
                'p5/12s36': 94.6667,
                'p5/12s37': 102,
                'p5/12s38': 102,
                'p5/12s39': 105.3333,
                'p5/12s40': 105.3333,
                'p6/12s10': 20,
                'p6/12s11': 23.3333,
                'p6/12s12': 23.3333,
                'p6/12s13': 26.6667,
                'p6/12s14': 27.3333,
                'p6/12s15': 32,
                'p6/12s16': 32,
                'p6/12s17': 42,
                'p6/12s18': 44.6667,
                'p6/12s19': 48,
                'p6/12s20': 48,
                'p6/12s21': 50,
                'p6/12s22': 55.3333,
                'p6/12s23': 57.3333,
                'p6/12s24': 57.3333,
                'p6/12s25': 62,
                'p6/12s26': 66.6667,
                'p6/12s27': 68.6667,
                'p6/12s28': 70,
                'p6/12s29': 73.3333,
                'p6/12s30': 74,
                'p6/12s31': 77.6667,
                'p6/12s32': 88,
                'p6/12s33': 90,
                'p6/12s34': 95.3333,
                'p6/12s35': 98.6667,
                'p6/12s36': 104,
                'p6/12s37': 106,
                'p6/12s38': 106,
                'p6/12s39': 110.6667,
                'p6/12s40': 116,
                'p7/12s10': 20,
                'p7/12s11': 23.3333,
                'p7/12s12': 25.3333,
                'p7/12s13': 27.3333,
                'p7/12s14': 30,
                'p7/12s15': 32,
                'p7/12s16': 32.6667,
                'p7/12s17': 46,
                'p7/12s18': 46,
                'p7/12s19': 48,
                'p7/12s20': 50.6667,
                'p7/12s21': 55.3333,
                'p7/12s22': 55.3333,
                'p7/12s23': 58.6667,
                'p7/12s24': 61.3333,
                'p7/12s25': 63.3333,
                'p7/12s26': 67.3333,
                'p7/12s27': 70.6667,
                'p7/12s28': 72,
                'p7/12s29': 76.6667,
                'p7/12s30': 82,
                'p7/12s31': 84.3333,
                'p7/12s32': 90.6667,
                'p7/12s33': 94,
                'p7/12s34': 102,
                'p7/12s35': 106.6667,
                'p7/12s36': 110.6667,
                'p7/12s37': 112.6667,
                'p7/12s38': 115.3333,
                'p7/12s39': 117.3333,
                'p7/12s40': 117.3333,
                'p8/12s10': 21.3333,
                'p8/12s11': 24,
                'p8/12s12': 25.3333,
                'p8/12s13': 30,
                'p8/12s14': 30.6667,
                'p8/12s15': 32.6667,
                'p8/12s16': 32.6667,
                'p8/12s17': 46,
                'p8/12s18': 48.6667,
                'p8/12s19': 50.6667,
                'p8/12s20': 53.3333,
                'p8/12s21': 56.6667,
                'p8/12s22': 56.6667,
                'p8/12s23': 65.3333,
                'p8/12s24': 65.3333,
                'p8/12s25': 67.3333,
                'p8/12s26': 71.3333,
                'p8/12s27': 74.6667,
                'p8/12s28': 77.3333,
                'p8/12s29': 82,
                'p8/12s30': 83.3333,
                'p8/12s31': 85.6667,
                'p8/12s32': 97.3333,
                'p8/12s33': 102,
                'p8/12s34': 108.6667,
                'p8/12s35': 112,
                'p8/12s36': 129.3333,
                'p8/12s37': 135.3333,
                'p8/12s38': 138,
                'p8/12s39': 141.3333,
                'p8/12s40': 145.3333,
                'p9/12s10': 22,
                'p9/12s11': 25.3333,
                'p9/12s12': 26,
                'p9/12s13': 30.6667,
                'p9/12s14': 30.6667,
                'p9/12s15': 33.3333,
                'p9/12s16': 36,
                'p9/12s17': 48.6667,
                'p9/12s18': 48.6667,
                'p9/12s19': 54.6667,
                'p9/12s20': 54.6667,
                'p9/12s21': 60.6667,
                'p9/12s22': 60.6667,
                'p9/12s23': 65.3333,
                'p9/12s24': 65.3333,
                'p9/12s25': 67.3333,
                'p9/12s26': 72.6667,
                'p9/12s27': 77.3333,
                'p9/12s28': 82.6667,
                'p9/12s29': 86,
                'p9/12s30': 86,
                'p9/12s31': 87,
                'p9/12s32': 102.6667,
                'p9/12s33': 111.3333,
                'p9/12s34': 114,
                'p9/12s35': 117.3333,
                'p9/12s36': 121.3333,
                'p9/12s37': 124.6667,
                'p9/12s38': 127.3334,
                'p9/12s39': 147.3334,
                'p9/12s40': 148,
                'p10/12s10': 22,
                'p10/12s11': 26,
                'p10/12s12': 28.6667,
                'p10/12s13': 30.6667,
                'p10/12s14': 31.3333,
                'p10/12s15': 36,
                'p10/12s16': 36.6667,
                'p10/12s17': 50,
                'p10/12s18': 50,
                'p10/12s19': 58.6667,
                'p10/12s20': 58.6667,
                'p10/12s21': 60.6667,
                'p10/12s22': 63.3333,
                'p10/12s23': 65.3333,
                'p10/12s24': 69.3333,
                'p10/12s25': 72.6667,
                'p10/12s26': 78,
                'p10/12s27': 81.3333,
                'p10/12s28': 84,
                'p10/12s29': 106.6667,
                'p10/12s30': 109.3333,
                'p10/12s31': 112.6667,
                'p10/12s32': 115.3334,
                'p10/12s33': 121.3334,
                'p10/12s34': 122.6666,
                'p10/12s35': 127.3333,
                'p10/12s36': 148.6667,
                'p10/12s37': 150.6667,
                'p10/12s38': 136,
                'p10/12s39': 159.3333,
                'p10/12s40': 164,
                'p11/12s10': 24,
                'p11/12s11': 26,
                'p11/12s12': 29.3333,
                'p11/12s13': 31.3333,
                'p11/12s14': 31.3333,
                'p11/12s15': 36.6667,
                'p11/12s16': 44.6667,
                'p11/12s17': 50,
                'p11/12s18': 56.6667,
                'p11/12s19': 58.6667,
                'p11/12s20': 58.6667,
                'p11/12s21': 63.3333,
                'p11/12s22': 67.3333,
                'p11/12s23': 69.3333,
                'p11/12s24': 70.6667,
                'p11/12s25': 80.6667,
                'p11/12s26': 79.3333,
                'p11/12s27': 102.6666,
                'p11/12s28': 105.3333,
                'p11/12s29': 108.6667,
                'p11/12s30': 111.3333,
                'p11/12s31': 117.3333,
                'p11/12s32': 118.6667,
                'p11/12s33': 138,
                'p11/12s34': 141.3334,
                'p11/12s35': 150,
                'p11/12s36': 153.3334,
                'p11/12s37': 159.3334,
                'p11/12s38': 161.3333,
                'p11/12s39': 164.6667,
                'p11/12s40': 169.3333,
                'p12/12s10': 24,
                'p12/12s11': 29.3333,
                'p12/12s12': 29.3333,
                'p12/12s13': 32,
                'p12/12s14': 34.6667,
                'p12/12s15': 37.3333,
                'p12/12s16': 45.3333,
                'p12/12s17': 56.6667,
                'p12/12s18': 56.6667,
                'p12/12s19': 58.6667,
                'p12/12s20': 65.3333,
                'p12/12s21': 67.3333,
                'p12/12s22': 67.3333,
                'p12/12s23': 70.6667,
                'p12/12s24': 77.3333,
                'p12/12s25': 96,
                'p12/12s26': 98.6666,
                'p12/12s27': 102,
                'p12/12s28': 106,
                'p12/12s29': 110.6667,
                'p12/12s30': 113.3333,
                'p12/12s31': 130.3334,
                'p12/12s32': 134.6667,
                'p12/12s33': 145.3334,
                'p12/12s34': 146.6667,
                'p12/12s35': 152.3333,
                'p12/12s36': 154.6666,
                'p12/12s37': 158.6666,
                'p12/12s38': 162.6667,
                'p12/12s39': 184,
                'p12/12s40': 202.6666
            }

        bf = bfCalcDict[bfCode]
        totalBf = float( bf ) * float( trussNumber )
        gables = float(gable) * (float(bf)+10)
        totalBf += gables
        price = float( totalBf ) * float( pPerBf )
        self.lPrice.setText( str( price ) )
        self.lTotalBf.setText( str( totalBf ) + ' BF' )

        print( bf )


    def tax(self):

        quotedPrice = self.inputQuotedPrice.text()

        if self.checkBoxCusTax.isChecked():
            tax = float( self.inputRegionCusTax.text() )
            quotedPrice = re.sub( '[!@#$,]', '', quotedPrice )
            quotedPrice = float( quotedPrice )
            taxCal = tax / 100
            totalWithTax = (quotedPrice * taxCal) + quotedPrice
            self.taxRate.setText( 'Tax Rate: ' + str( tax ) + '%' )
            self.inputTotalPrice.setText( str( totalWithTax ) )
            zipCode = self.inputZipCode.text()


            search = SearchEngine( simple_zipcode=True )
            zipcode = search.by_zipcode( zipCode )


            zipcodecity = zipcode.common_city_list


            self.shippingCity.clear()



            self.shippingCity.addItems( zipcodecity )

            zip_region = zipcode.state
            self.inputRegion.setText(zip_region)

        if self.checkBoxCusTax.isChecked() == False:

            print( 'start' )



            zipCode = self.inputZipCode.text()


            search = SearchEngine( simple_zipcode=True )
            zipcode = search.by_zipcode( zipCode )


            zipcodecity = zipcode.common_city_list


            self.shippingCity.clear()


            self.shippingCity.addItems( zipcodecity )

            zip_region = zipcode.state
            self.inputRegion.setText( zip_region )


            quotedPrice = re.sub( '[!@#$,]', '', quotedPrice )
            quotedPrice = float( quotedPrice )
            totalWithTax = quotedPrice




        if self.taxFree.isChecked():
            totalWithTax = quotedPrice

        self.inputTotalPrice.setText( str( totalWithTax ) )

    def quoteForm(self):

        try:

            jobNumber = self.inputJobNumber.currentText()

            o = win32com.client.Dispatch( "Excel.Application" )

            o.Visible = False

            wb_path = os.getcwd() + '\\' + jobNumber + '.xlsx'

            wb = o.Workbooks.Open( wb_path )

            ws_index_list = [2]  # say you want to print these sheets

            path_to_pdf = os.getcwd() + '\\' + jobNumber + ' Quote.pdf'

            print_area = 'A1:K42'

            for index in ws_index_list:

                # off-by-one so the user can start numbering the worksheets at 1

                ws = wb.Worksheets[index - 1]

                ws.PageSetup.Zoom = False

                ws.PageSetup.FitToPagesTall = 1

                ws.PageSetup.FitToPagesWide = 1

                ws.PageSetup.PrintArea = print_area

            wb.WorkSheets( ws_index_list ).Select()

            wb.ActiveSheet.ExportAsFixedFormat( 0, path_to_pdf )
            wb.Close( True )
            pyautogui.alert( 'Quote Form Created' )
            if not os.path.exists( r'O:\Jobs\\' + jobNumber + '\Quote' ):
                os.mkdir( r'O:\Jobs\\' + jobNumber + '\Quote' )

            file = jobNumber + ' Quote.pdf'
            shutil.move( os.getcwd() + '\\' + file, 'O:\Jobs\\' + jobNumber + '\Quote\\' + file )

            pyautogui.alert( 'did it work?' )
        except:
            pyautogui.alert( 'Unable to creatE pdf Have you Consolidated this job yet?' )

    def jobManagment(self):

        jobNameEst = self.jobNameEst.text()
        priceEst = self.lPrice.text()
        bfEst = lTotalBf = self.lTotalBf.text()

        runNameEst = self.runNameEst.text()

        if not os.path.exists( jobNameEst ):
            os.mkdir( jobNameEst )

        file = jobNameEst + runNameEst
        jobEstInfo = {'priceEst': priceEst, 'bfEst': bfEst}
        pickle.dump( jobEstInfo, open( file, "wb" ) )

        shutil.move( os.getcwd() + '\\' + file, os.getcwd() + '\\' + jobNameEst + '\\' + file )

        rowPosition = self.tableJobEst.rowCount()
        self.tableJobEst.insertRow( rowPosition )
        self.tableJobEst.setItem( rowPosition, 0, PySide2.QtWidgets.QTableWidgetItem( jobNameEst + ' ' + runNameEst ) )
        self.tableJobEst.setItem( rowPosition, 1, PySide2.QtWidgets.QTableWidgetItem( bfEst ) )
        self.tableJobEst.setItem( rowPosition, 2, PySide2.QtWidgets.QTableWidgetItem( priceEst ) )

    def jobTotal(self):
        val = sum( [float( item.text() ) for item in self.tableJobEst.selectedItems()] )
        table = PySide2.QtWidgets.QTableWidgetItem()
        # table.setText(str(val))
        val = round( val, 2 )
        # row = self.tableJobEst.currentRow()
        # self.tableJobEst.setItem(row, 2, table)
        rowPosition = self.tableJobEst.rowCount()
        self.tableJobEst.insertRow( rowPosition )
        self.tableJobEst.setItem( rowPosition, 0, PySide2.QtWidgets.QTableWidgetItem( '0' ) )
        self.tableJobEst.setItem( rowPosition, 1, PySide2.QtWidgets.QTableWidgetItem( '0' ) )
        self.tableJobEst.setItem( rowPosition, 2, PySide2.QtWidgets.QTableWidgetItem( '$' + str( val ) ) )

    def findJob(self):
        self.tableJobEst.setRowCount( 0 )
        jobNameEst = self.jobNameEst.text()
        if os.path.exists( jobNameEst ):
            path = jobNameEst

            files = []
            # r=root, d=directories, f = files
            for r, d, f in os.walk( path ):
                for file in f:
                    files.append( os.path.join( file ) )

            for f in files:
                jobInfo = pickle.load( open( path + '\\' + f, "rb" ) )
                priceEst = jobInfo.get( 'priceEst', '' )
                bfEst = jobInfo.get( 'bfEst', '' )

                rowPosition = self.tableJobEst.rowCount()
                self.tableJobEst.insertRow( rowPosition )
                self.tableJobEst.setItem( rowPosition, 0, PySide2.QtWidgets.QTableWidgetItem( f ) )
                self.tableJobEst.setItem( rowPosition, 1, PySide2.QtWidgets.QTableWidgetItem( bfEst ) )
                self.tableJobEst.setItem( rowPosition, 2, PySide2.QtWidgets.QTableWidgetItem( priceEst ) )

    def deleteSel(self):
        jobNameEst = self.jobNameEst.text()
        path = jobNameEst
        selected = self.tableJobEst.currentRow()
        fileName = self.tableJobEst.item( selected, 0 )
        self.ID = fileName.text()
        os.remove( path + '\\' + self.ID )

        self.tableJobEst.removeRow( selected )

    def bfEstFloor(self):
        depth = self.depthFloor.currentText()
        span = self.spanFloor.value()
        trussNumber = self.trussNumberFloor.text()

        if depth == '12" - 16"':
            price = span * 2.75

        if depth == '18" - 24"':
            price = span * 3.20

        self.lPrice.setText( str( price ) )

    def openXl(self):

        jobNumber = self.inputJobNumber.currentText()

        o = win32com.client.Dispatch( "Excel.Application" )

        o.Visible = True

        wb_path = os.getcwd() + '\\xldocs\\' + jobNumber + '.xlsx'

        wb = o.Workbooks.Open( wb_path )

    def openJobFolder(self):
        jobNumber = self.inputJobNumber.currentText()

        path = 'O:\Jobs\\' + jobNumber
        pyperclip.copy( path )

    def copyjob(self):

        jobNumber = pyautogui.prompt( 'job number' )
        jobName = self.inputJobName.text()
        salesman = self.inputSalesman.currentText()
        designer = self.inputDesigner.currentText()
        region = self.inputRegion.text()
        street = self.inputStreetName.text()
        zipCode = self.inputZipCode.text()
        quotedPrice = self.inputQuotedPrice.text()
        totalPrice = self.inputTotalPrice.text()
        customerCode = self.inputCustomerCode.text()
        customerName = self.inputCustomerName.text()
        billingStreet = self.inputBillingStreet.text()
        billingCity = self.inputBillingCity.text()
        billingZip = self.inputBillingZip.text()
        totalBf = self.inputBf.text()
        phoneNumber = self.inputPhoneNumber.text()
        email = self.inputEmail.text()
        date = self.inputDate.text()
        po = self.inputPo.text()

        jobInfo = {'jobnumber': jobNumber, 'Job Name': jobName, 'Salesman': salesman, 'Designer': designer,
                   'Region': region, 'Street Name': street, 'Zip Code': zipCode, 'Quoted Price': quotedPrice,
                   'Total Price': totalPrice, 'Customer Code': customerCode, 'Customer Name': customerName,
                   'Billing Street': billingStreet, 'Billing City': billingCity, 'Billing Zip': billingZip,
                   'Bf': totalBf, 'phonenumber': phoneNumber, 'email': email, 'date': date, 'po': po}
        pickle.dump( jobInfo, open( jobNumber, "wb" ) )

        file = jobNumber
        shutil.move( os.getcwd() + '\\' + file, os.getcwd() + '\\Jobs\\' + file )

    def custLookup(self):
        comboCust = self.window.findChild( QComboBox, 'comboCust' )
        cust = comboCust.currentText()
        try:
            custInfo = pickle.load( open( 'Cust\\' + cust, "rb" ) )
            customerCode = custInfo.get( 'Customer Code', '' )
            customerName = custInfo.get( 'Customer Name', '' )
            billingStreet = custInfo.get( 'Billing Street', '' )
            billingCity = custInfo.get( 'Billing City', '' )
            billingZip = custInfo.get( 'Billing Zip', '' )
            phoneNumber = custInfo.get( 'phonenumber', '' )
            email = custInfo.get( 'email', '' )

            self.inputCustomerCode.setText( customerCode )
            self.inputCustomerName.setText( customerName )
            self.inputBillingStreet.setText( billingStreet )
            self.inputBillingCity.setText( billingCity )
            self.inputBillingZip.setText( billingZip )
            self.inputPhoneNumber.setText( phoneNumber )
            self.inputEmail.setText( email )
        except:
            print( 'nocust' )
            # self.load()

    def saveCustAct(self):
        customerCode = self.inputCustomerCode.text()
        customerName = self.inputCustomerName.text()
        billingStreet = self.inputBillingStreet.text()
        billingCity = self.inputBillingCity.text()
        billingZip = self.inputBillingZip.text()
        phoneNumber = self.inputPhoneNumber.text()
        email = self.inputEmail.text()

        custInfo = {'Customer Code': customerCode, 'Customer Name': customerName, 'Billing Street': billingStreet,
                    'Billing City': billingCity, 'Billing Zip': billingZip, 'phonenumber': phoneNumber, 'email': email}
        pickle.dump( custInfo, open( customerCode, "wb" ) )

        file = customerCode
        shutil.move( os.getcwd() + '\\' + file, os.getcwd() + '\\Cust\\' + file )

        comboCust = self.window.findChild( QComboBox, 'comboCust' )

        comboCust.addItem( file )

    def delete(self):
        confirm = pyautogui.confirm( 'Are you sure you want to delete this job?' )
        print( confirm )
        if confirm == 'OK':
            job = self.inputJobNumber.currentText()
            file = os.getcwd() + '\\Jobs\\' + job
            os.remove( file )

    def load_project_structure(self, startpath, tree):
        jobNumber = self.inputJobNumber.currentText()

        for element in os.listdir( startpath ):
            path_info = startpath + "/" + element
            parent_itm = QTreeWidgetItem( tree, [os.path.basename( element )] )
            if os.path.isdir( path_info ):
                self.load_project_structure( path_info, parent_itm )

            else:
                print( 'no' )

    def openfile(self):

        jobNumber = self.inputJobNumber.currentText()

        path = 'O:\Jobs\\' + jobNumber

        item = self.treeWidget.currentItem()

        col = self.treeWidget.currentColumn()

        itemtext = item.text( col )
        a = self.treeWidget.indexOfTopLevelItem( self.treeWidget.currentItem() )

        if self.treeWidget.indexOfTopLevelItem( self.treeWidget.currentItem() ) == -1:
            parent1 = self.treeWidget.currentItem().parent()
            if self.treeWidget.indexOfTopLevelItem( parent1 ) == -1:
                parent2 = parent1.parent()
                parent1text = parent1.text( col )
                parent2text = parent2.text( col )

                toOpen = path + '\\' + parent2text + '\\' + parent1text + '\\' + itemtext

                pyautogui.alert( toOpen )
                os.startfile( toOpen )

                return

            parent1text = parent1.text( col )

            toOpen = path + '\\' + parent1text + '\\' + itemtext
            pyautogui.alert( toOpen )
            os.startfile( toOpen )
        else:
            toOpen = path + '\\' + itemtext
            pyautogui.alert( toOpen )
            os.startfile( toOpen )

        # itemtext=item.text()

    def addSalesmen(self):
        newSalesmen = pyautogui.prompt( 'Salesman to add' )
        wd = os.getcwd()
        salesman = pickle.load( open( wd + '\\Salesmen\\sales', "rb" ) )
        salesmanList = salesman.get( 'salesmen' )
        salesmanList.append( str( newSalesmen ) )
        salesman = {'salesmen': salesmanList}
        pickle.dump( salesman, open( wd + '\\Salesmen\\sales', "wb" ) )
        salesman = pickle.load( open( wd + '\\Salesmen\\sales', "rb" ) )
        self.inputSalesman.clear()
        salesmanList = salesman.get( 'salesmen' )
        self.inputSalesman.addItems( salesmanList )
        print( salesman )

    def removeSalesmen(self):
        print( "Opening a new popup window..." )
        self.w = MyPopup()
        self.w.setGeometry( QRect( 100, 100, 400, 200 ) )
        self.w.show()

    def reloadprog(self):
        wd = os.getcwd()
        salesman = pickle.load( open( wd + '\\Salesmen\\sales', "rb" ) )
        self.inputSalesman.clear()
        salesmanList = salesman.get( 'salesmen' )
        self.inputSalesman.addItems( salesmanList )

    def search(self):
        self.w = MySearch()
        self.w.setGeometry( QRect( 100, 100, 400, 200 ) )
        self.w.show()




# TODO Feature : Archive
# TODO Feature : Drag and drop files? 


# TODO Game Changer : Scan PDFS for info on jobs E.G train python to look at a shipping ticket or quote and grab info from it like notes, dates, ect..


if __name__ == '__main__':
    app = QApplication( sys.argv )
    form = Form( 'mainwindow.ui' )
    sys.exit( app.exec_() )
